import openpyxl
import re
import sys
import os
from copy import copy
from datetime import datetime
from openpyxl.utils import get_column_letter, column_index_from_string


# ── helpers ───────────────────────────────────────────────────────────────────

def read_header(ws):
    account_str = date_range = downloaded = ""
    for row in ws.iter_rows(min_row=1, max_row=4, values_only=True):
        for cell in row:
            if cell and isinstance(cell, str):
                if "Account:" in cell:
                    account_str = cell
                elif "Date Range:" in cell:
                    date_range = cell.replace("Date Range: ", "").strip()
                elif "Downloaded:" in cell:
                    downloaded = cell.replace("Downloaded: ", "").strip()
    return account_str, date_range, downloaded


def find_header_row(ws, max_scan=10):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), 1):
        non_empty = [c for c in row if c is not None]
        if len(non_empty) > 3:
            return i
    return None


def tab_to_dict(ws):
    header_row = find_header_row(ws)
    if header_row is None:
        return {}
    rows = list(ws.iter_rows(min_row=header_row, max_row=header_row + 1, values_only=True))
    if len(rows) < 2:
        return {}
    headers, data_row = rows[0], rows[1]
    result = {}
    for i, h in enumerate(headers):
        if h is not None and i < len(data_row):
            result[h] = data_row[i]
    return result


def tab_to_records(ws):
    header_row = find_header_row(ws)
    if header_row is None:
        return []
    headers = None
    records = []
    for i, row in enumerate(ws.iter_rows(min_row=header_row, values_only=True), header_row):
        if headers is None:
            headers = list(row)
            continue
        if not any(row):
            continue
        rec = {headers[j]: row[j] for j in range(len(headers)) if headers[j] is not None}
        records.append(rec)
    return records


def safe(val, default=""):
    return default if val is None else val


def _adjust_formula(formula: str, from_row: int, to_row: int) -> str:
    """Adjust row numbers in a formula string."""
    def replace_ref(m):
        col_part = m.group(1)
        row_num = int(m.group(2))
        delta = row_num - from_row
        return f"{col_part}{to_row + delta}"
    return re.sub(r'([A-Z]+)(\d+)', replace_ref, formula)


def _copy_cell_format(src_cell, dst_cell):
    """Copy number format and font from src to dst (no fill copy — keep theme)."""
    if src_cell.number_format:
        dst_cell.number_format = src_cell.number_format
    if src_cell.font:
        dst_cell.font = copy(src_cell.font)
    if src_cell.alignment:
        dst_cell.alignment = copy(src_cell.alignment)


def _clear_cell(cell):
    """Fully blank a cell — value, number format, font."""
    cell.value = None
    cell.number_format = "General"


# ── main writer ───────────────────────────────────────────────────────────────

def write_strategy(pre_analysis_path: str, template_path: str, output_dir: str):

    pa = openpyxl.load_workbook(pre_analysis_path, data_only=True, read_only=True)

    # ── header ────────────────────────────────────────────────────────────────
    ws01 = pa['01_Advertiser_Name']
    account_str, date_range, downloaded = read_header(ws01)

    m = re.match(r"Account:\s*(.+?)\s*\|\s*Tenant ID:\s*(\S+)\s*\|\s*Account ID:\s*(\S+)", account_str)
    if not m:
        raise ValueError(f"Could not parse account string: {account_str}")
    account_label = m.group(1).strip()
    tenant_id     = m.group(2).strip()
    profile_id    = m.group(3).strip()
    member_id     = account_label.split(" - ")[0].strip()

    # ── tab 55 ────────────────────────────────────────────────────────────────
    d55 = tab_to_dict(pa['55_Salesforce_Consolidated_PreA'])

    # ── tab 38 ───────────────────────────────────────────────────────────────
    d38 = tab_to_dict(pa['38_Client_Success_Insights_Repo'])

    # ── tab 37 gong ──────────────────────────────────────────────────────────
    gong_records = tab_to_records(pa['37_Gong_Call_Insights_for_Sales'])
    gong = gong_records[0] if gong_records else {}

    # ── tab 14 ASIN data — optional (vendor accounts may not have this tab) ──
    tab14_name = '14_Campaign_Performance_by_Adve'
    if tab14_name in pa.sheetnames:
        asin_records = tab_to_records(pa[tab14_name])
        vendor_account = False
    else:
        asin_records = []
        vendor_account = True

    # ── tab 54 ────────────────────────────────────────────────────────────────
    d54 = tab_to_dict(pa['54_Project_Dataset_on_SF'])

    # ── tab 22 catalogue ─────────────────────────────────────────────────────
    cat_records = tab_to_records(pa['22_Catalogue_Details'])
    cat_by_asin = {r['asin']: r for r in cat_records if r.get('asin')}

    pa.close()

    # ── load template ─────────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(template_path, keep_vba=True)

    # ════════════════════════════════════════════════════════════════════════════
    # TAB 1 — Questionaire Survey - AMZ
    # ════════════════════════════════════════════════════════════════════════════
    ws1 = wb['Questionaire Survey - AMZ']

    def w1(coord, value):
        ws1[coord] = value

    w1('C6', member_id)
    w1('F6', profile_id)
    w1('J6', safe(d55.get('CSP_Last_Modified_By')))

    w1('F7', profile_id)
    w1('J7', safe(d55.get('Projected_Project_MRR__c')))

    w1('C8', safe(d55.get('Account_Name')))
    ld = d55.get('Launch_Date__c')
    w1('F8', ld.strftime('%Y-%m-%d') if hasattr(ld, 'strftime') else safe(ld))
    if hasattr(ld, 'strftime'):
        months = (datetime.now().year - ld.year) * 12 + (datetime.now().month - ld.month)
        w1('J8', f"{months} months")
    else:
        w1('J8', safe(d38.get('Customer_Age_Months__c')))

    w1('C9', safe(d55.get('Customer_Age_Months__c') or d38.get('Customer_Age_Months__c')))
    w1('F9', safe(d38.get('Repeat_Purchase_Behavior__c')))
    w1('J9', safe(d55.get('CSM_Churn_Risk__c')))

    w1('C10', safe(d38.get('Commodity_Products_or_Branded_Products__c')))
    w1('F10', safe(d38.get('Sales_Concentration__c')))
    w1('J10', safe(d55.get('Director_Churn_Risk__c')))

    w1('C11', safe(d55.get('CSM')))
    w1('F11', safe(d38.get('CSM_Tenure__c')))
    w1('J11', safe(d55.get('Account_Risk_Score__c')))

    w1('C12', '')
    w1('F12', safe(d55.get('Active_Products__c')))

    w1('F13', safe(d38.get('Customer_Feedback__c')))

    w1('C15', safe(d55.get('Current_Challenges__c')))
    w1('F15', safe(d55.get('Primary_Objective__c')))
    w1('J15', safe(d55.get('ACOS_Constraint__c')))

    w1('C16', safe(d55.get('Primary_Objective_Additional_Context__c')))
    w1('F16', safe(d55.get('Primary_Spend_KPI__c')))
    w1('J16', safe(d38.get('Customer_Acquisition_Cost_Target__c')))

    w1('C17', safe(d55.get('Top_Priority__c')))
    w1('J17', safe(d55.get('TACOS_Constraint__c')))

    w1('C18', safe(d55.get('Second_Priority__c')))
    w1('J18', safe(d55.get('daily_target_spend__c')))
    w1('F18', safe(d54.get('CS_Notes__c')))

    w1('C19', safe(d55.get('Biggest_Expansion_Opportunity__c')))
    w1('F19', safe(d55.get('Near_Term_3_Month_Considerations__c')))
    w1('J19', safe(d55.get('Target_ROAS__c')))

    stage_rows = {1: (24, 25), 2: (27, 28), 3: (30, 31), 4: (33, 34)}
    for s, (r_a, r_i) in stage_rows.items():
        w1(f'C{r_a}', safe(d55.get(f'AdoptionOrUpsellS{s}__c')))
        w1(f'G{r_a}', safe(d55.get(f'StrategyS{s}__c')))
        w1(f'J{r_a}', safe(d55.get(f'StatusS{s}__c')))
        intro = d55.get(f'ExecutionDateS{s}__c')
        w1(f'C{r_i}', intro.strftime('%Y-%m-%d') if hasattr(intro, 'strftime') else safe(intro))

    w1('C41', safe(gong.get('Gong__Call_Brief__c') or d55.get('Call_Brief')))
    w1('C42', safe(gong.get('Gong__Call_Key_Points__c') or d55.get('Key_Points')))
    w1('C43', safe(gong.get('Gong__Call_Highlights_Next_Steps__c') or d55.get('Highlights_Next_Steps')))

    # ════════════════════════════════════════════════════════════════════════════
    # TAB 2 — Account Strategy _Analysis header
    # ════════════════════════════════════════════════════════════════════════════
    ws2 = wb['Account Strategy _Analysis']
    ws2['A1'] = f"{account_label} — Account Strategy Analysis"
    ws2['B3'] = f"Account: {account_label} | Tenant ID: {tenant_id} | Account ID: {profile_id}"
    ws2['B4'] = date_range
    ws2['B5'] = downloaded

    # ════════════════════════════════════════════════════════════════════════════
    # TAB 3 — ChildASIN View
    # ════════════════════════════════════════════════════════════════════════════
    ws3 = wb['ChildASIN View']

    col14_map = {
        'Parent ASIN':        'ParentASIN',
        'ASIN':               'asin',
        'Total Sales':        'TotalSales',
        'Total Units Ordered':'UnitsOrdered',
        'Ad Spend':           'AdSpend',
        'TACoS':              'TACoS',
        'Ad Sales':           'AdSales',
        'Ads Units Ordered':  'Orders',
        'ACoS':               'ACoS',
        'Clicks':             'Clicks',
        'Tier':               'Tier',
        'Buy Box%':           'Weighted_BuyBoxPercentage',
        'ATM_Spend':          'ATM_Spend',
        'BA_Spend':           'BA_Spend',
        'Manual_Q1_Spend':    'Manual_Q1_Spend',
        'BAK_Spend':          'BAK_Spend',
        'OP_Spend':           'OP_Spend',
        'SPT_Spend':          'SPT_Spend',
        'CAT_SP_Spend':       'CAT_SP_Spend',
        'WATM_Spend':         'WATM_Spend',
        'SB_Spend':           'SB_Spend',
        'SBV_Spend':          'SBV_Spend',
        'SD_Spend':           'SD_Spend',
        'Imported_Spend':     'Imported_Spend',
        'NonQuartile_Spend':  'NonQuartile_Spend',
        'TAG 1':              'Tag1',
        'TAG 2':              'Tag2',
        'TAG 3':              'Tag3',
        'TAG 4':              'Tag4',
        'TAG 5':              'Tag5',
    }

    col22_map = {
        'AOV':        'AOV',
        'PriceTier':  'PriceTier',
        'Brand':      'Brand',
        'Department': 'Department',
        'Category':   'Category',
    }

    # Columns K-AM that have formulas in row 3 (template reference row)
    FORMULA_COLS = {
        11: '=IFERROR(IF({col}{row}/C{row} <= 0.3, 30, IF({col}{row}/C{row} <= 0.5, 50, IF({col}{row}/C{row} <= 0.8, 80, 100))),"-")',
        12: '=IFERROR(G{row}/C{row},"-")',
        13: '=IFERROR((C{row}-G{row})/C{row},"-")',
        28: '=SUM(O{row},Q{row},S{row})',
        29: '=SUM(P{row},R{row},T{row},V{row},Z{row},AA{row},W{row},X{row},Y{row})',
    }

    # Capture number formats from template row 3 for cols K-AM
    K_COL = column_index_from_string('K')
    AM_COL = column_index_from_string('AM')
    col_numfmt = {}
    for col in range(K_COL, AM_COL + 1):
        col_numfmt[col] = ws3.cell(3, col).number_format

    # Build header → column index from row 2
    header_to_col = {}
    for cell in ws3[2]:
        if cell.value:
            header_to_col[cell.value] = cell.column

    # ── Vendor account: write "No Data" notice and clear the rest ────────────
    if vendor_account:
        # Clear all data rows
        for row in ws3.iter_rows(min_row=3, max_col=ws3.max_column):
            for cell in row:
                _clear_cell(cell)
        # Write a single "No Data" message
        ws3.cell(row=3, column=1).value = "No Data — Vendor account (tab 14 not available)"
        # Save and return early
        filename = f"{account_label} — Strategy Analysis {date_range}.xlsm"
        filename = re.sub(r'[<>:"/\\|?*]', '-', filename)
        out_path = os.path.join(output_dir, filename)
        wb.save(out_path)
        print(f"Saved (vendor, no ASIN data): {out_path}")
        return out_path

    # ── Clear ALL existing data and formatting from row 3 down ───────────────
    for row in ws3.iter_rows(min_row=3, max_col=ws3.max_column):
        for cell in row:
            _clear_cell(cell)

    # ── Write data rows — only for actual records ────────────────────────────
    total_sales_all = (asin_records[0].get('TotalSalesAll') or 1) if asin_records else 1

    for row_idx, rec in enumerate(asin_records, start=3):
        asin = rec.get('asin', '')
        cat  = cat_by_asin.get(asin, {})

        # Write raw data cols (A-J and O-AM writer-owned columns)
        for header, col_idx in header_to_col.items():
            val = None

            if header in col14_map:
                val = rec.get(col14_map[header])
            elif header == 'Ad Sales (%)':
                # Handled by formula in col L — skip raw write
                continue
            elif header == 'Organic Sales (%)':
                # Handled by formula in col M — skip raw write
                continue
            elif header in col22_map:
                val = cat.get(col22_map[header])

            if val is not None and col_idx not in FORMULA_COLS:
                ws3.cell(row=row_idx, column=col_idx, value=val)

        # Apply formulas for cols K, L, M, AB, AC
        for col_idx, formula_tpl in FORMULA_COLS.items():
            col_ltr = get_column_letter(col_idx)
            if '{col}' in formula_tpl:
                formula = formula_tpl.format(col='E', row=row_idx)
            else:
                formula = formula_tpl.format(row=row_idx)
            ws3.cell(row=row_idx, column=col_idx, value=formula)

        # Apply number formats cols K-AM for this data row only
        for col in range(K_COL, AM_COL + 1):
            cell = ws3.cell(row=row_idx, column=col)
            fmt = col_numfmt.get(col, 'General')
            if fmt and fmt != 'General':
                cell.number_format = fmt

    # ── save ──────────────────────────────────────────────────────────────────
    filename = f"{account_label} — Strategy Analysis {date_range}.xlsm"
    filename = re.sub(r'[<>:"/\\|?*]', '-', filename)
    out_path = os.path.join(output_dir, filename)
    wb.save(out_path)
    print(f"Saved: {out_path}")
    return out_path


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python writer_strategy.py <pre_analysis.xlsx> <template.xlsm> [output_dir]")
        sys.exit(1)
    write_strategy(sys.argv[1], sys.argv[2], sys.argv[3] if len(sys.argv) > 3 else "/mnt/user-data/outputs")
